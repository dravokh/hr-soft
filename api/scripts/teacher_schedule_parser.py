#!/usr/bin/env python3
"""
Helper CLI that mirrors the PHP-based TeacherScheduleAnalyzer logic.
The script is optional and is mainly intended for debugging locally.
PDF parsing requires the `pypdf` package and XLSX parsing uses `openpyxl`.
Install them with: pip install pypdf openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover - dependency is optional
    PdfReader = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - dependency is optional
    load_workbook = None


CAMBRIDGE_RE = re.compile(r"\b\d+C-A\b", re.IGNORECASE)
GEORGIAN_RE = re.compile(r"\b\d+-A\b", re.IGNORECASE)


def _flatten_lines(lines: Iterable[str]) -> List[str]:
    sanitized: List[str] = []
    for line in lines:
        clean = re.sub(r"\s+", " ", line).strip()
        if clean:
            sanitized.append(clean)
    return sanitized


def _build_payload(text: str, fallback: str) -> dict:
    lines = _flatten_lines(text.splitlines())
    teacher = next(iter(lines), fallback or "Unknown")
    cambridge = len(CAMBRIDGE_RE.findall(text))
    georgian = len(GEORGIAN_RE.findall(text))
    return {
        "teacher": teacher,
        "cambridgeCount": cambridge,
        "georgianCount": georgian,
    }


def _parse_pdf(path: Path) -> List[Tuple[str, str]]:
    if PdfReader is None:
        raise RuntimeError("pypdf is required to parse PDF files. Run `pip install pypdf`.")
    reader = PdfReader(str(path))
    pages: List[Tuple[str, str]] = []
    for idx, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            pages.append((text, f"Page {idx}"))
    return pages


def _parse_xlsx(path: Path) -> List[Tuple[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse XLSX files. Run `pip install openpyxl`.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    pages: List[Tuple[str, str]] = []
    for sheet in workbook.worksheets:
        lines: List[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells: List[str] = []
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    cells.append(text)
            if cells:
                lines.append(" ".join(cells))
        if lines:
            pages.append(("\n".join(lines), sheet.title or "Sheet"))
    return pages


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Analyze teacher schedule files.")
    parser.add_argument("file", type=Path, help="Path to the PDF or XLSX input file")
    args = parser.parse_args(argv)

    file_path = args.file
    if not file_path.exists():
        parser.error(f"{file_path} does not exist.")

    extension = file_path.suffix.lower().lstrip(".")
    if extension == "pdf":
        pages = _parse_pdf(file_path)
    elif extension == "xlsx":
        pages = _parse_xlsx(file_path)
    else:
        parser.error("Only PDF or XLSX files are supported.")
        return 2

    teachers = [_build_payload(text, label) for text, label in pages]
    json.dump({"teachers": teachers}, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

